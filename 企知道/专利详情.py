import requests
from Crypto.Cipher import AES
import base64
import json
from datetime import datetime
from openpyxl import Workbook

headers = {
    "accept": "application/json, text/plain, */*",
    "accept-language": "zh-CN,zh;q=0.9",
    "accesstoken": "eyJhbGciOiJIUzUxMiJ9.ZXlKNmFYQWlPaUpFUlVZaUxDSmhiR2NpT2lKa2FYSWlMQ0psYm1NaU9pSkJNVEk0UTBKRExVaFRNalUySW4wLi5jUmtMZndKWGhISUdpdW5DM0x4Z2ZRLjYtbmpGQXJBSDhuM01Zcy0zVkJyNGx6NTRsWjNkMTZiNW5objNsaFZoRV9lU0JnU2YzaGlRWDduTFh3cEtJUnhmMXZybGtSRmVPNGxnZ1pzazRRM3JXNjlfbndUUjdBdTdoeUh4NVZRTkdmWG9VVFRmT0k4NGJfaWRzalBwaENyRVExNWdFNHR2TFVXTE5RQTJRUGw5c0lzT2haeW50bHJBTU15UTBZX1JRbG01cktMeEVGd2hMb0xjR0FmaV96dGJCNVJRbGdZSHVIYjFfWkd5OXJMNEVWVWoxeld0VER4SDFBV3laNmlfRGcuS1NzSndKbGo2YllHRFJTcjdQM3VGZw.mnYj8Fu3jzieBZ7Z2TBIDcvk98fJsUis60RSyrEFePKsRo6rz8MdDokcC8b1GeFMrEmyjvYaBuCRIfBzmkR8BA",
    "cache-control": "no-cache",
    "content-type": "application/json",
    "device-id": "BEQEYsd27lvqjrdA/NmjwqUGRIb+2PH/wkqKp8ozz3G6BYV4O2Z42OTBkWhIDJXM5A5KjO95aSNa0vvd1LA9YCg==",
    "h5version": "v1.0.0",
    "origin": "https://patents.qizhidao.com",
    "pragma": "no-cache",
    "priority": "u=1, i",
    "referer": "https://patents.qizhidao.com/search/detail/WZIP_14feea8dc3c20cb97f1de11e4233613c?orderColumn=undefined&searchType=simple_search&orderType=undefined&filter=&tab=0&from=result&businessSource=%E6%9F%A5%E4%B8%93%E5%88%A9-%E6%90%9C%E7%B4%A2%E7%BB%93%E6%9E%9C%E5%88%97%E8%A1%A8-%E6%90%9C%E7%B4%A2%E6%9B%B4%E5%A4%9A&statement=ASPS_EXACT%3A%28%22%E8%A5%BF%E5%8D%97%E8%B4%A2%E7%BB%8F%E5%A4%A7%E5%AD%A6%E5%A4%A9%E5%BA%9C%E5%AD%A6%E9%99%A2%22%29&semanticId=&patentName=%E4%BD%93%E8%82%B2%E5%99%A8%E6%9D%90%E8%87%AA%E5%8A%A8%E5%80%9F%E8%BF%98%E4%B8%80%E4%BD%93%E6%9C%BA&rightSidebar=false&imageSessionKey=&simpleMode=&proVersion=1&current=1&pageSize=20",
    "sec-ch-ua": "\"Chromium\";v=\"130\", \"Google Chrome\";v=\"130\", \"Not?A_Brand\";v=\"99\"",
    "sec-ch-ua-mobile": "?0",
    "sec-ch-ua-platform": "\"Windows\"",
    "sec-fetch-dest": "empty",
    "sec-fetch-mode": "cors",
    "sec-fetch-site": "same-site",
    "sensordeviceid": "1983f613e1a156a-0bae2da6cdc884-26011951-1327104-1983f613e1b1723",
    "sensorsdistinctid": "c0db24d8680b49b4a8678ad0a36910e8",
    "signature": "2fb2a6c09366772b87278e91fe3b48ee.z64U0z",
    "token": "eyJhbGciOiJIUzUxMiJ9.ZXlKNmFYQWlPaUpFUlVZaUxDSmhiR2NpT2lKa2FYSWlMQ0psYm1NaU9pSkJNVEk0UTBKRExVaFRNalUySW4wLi5jUmtMZndKWGhISUdpdW5DM0x4Z2ZRLjYtbmpGQXJBSDhuM01Zcy0zVkJyNGx6NTRsWjNkMTZiNW5objNsaFZoRV9lU0JnU2YzaGlRWDduTFh3cEtJUnhmMXZybGtSRmVPNGxnZ1pzazRRM3JXNjlfbndUUjdBdTdoeUh4NVZRTkdmWG9VVFRmT0k4NGJfaWRzalBwaENyRVExNWdFNHR2TFVXTE5RQTJRUGw5c0lzT2haeW50bHJBTU15UTBZX1JRbG01cktMeEVGd2hMb0xjR0FmaV96dGJCNVJRbGdZSHVIYjFfWkd5OXJMNEVWVWoxeld0VER4SDFBV3laNmlfRGcuS1NzSndKbGo2YllHRFJTcjdQM3VGZw.mnYj8Fu3jzieBZ7Z2TBIDcvk98fJsUis60RSyrEFePKsRo6rz8MdDokcC8b1GeFMrEmyjvYaBuCRIfBzmkR8BA",
    "traceparent": "00-3da10dbc93f540f091b13e145cfcba09-5143633543749831-01",
    "tracestate": "rum=v2&browser&fyw9n1jhpf@18ff4c0a0d8fd8d&819ce306a1f84463a20db7cdd174591f&uid_tfxjlwhu0pc63iii",
    "user-agent": "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/130.0.0.0 Safari/537.36",
    "user-agent-web": "X/32ed80a37cf74a5697657feae2d62323"
}
cookies = {
    "sajssdk_2015_cross_new_user": "1",
    "wz_uuid": "X/32ed80a37cf74a5697657feae2d62323",
    "token": "eyJhbGciOiJIUzUxMiJ9.ZXlKNmFYQWlPaUpFUlVZaUxDSmhiR2NpT2lKa2FYSWlMQ0psYm1NaU9pSkJNVEk0UTBKRExVaFRNalUySW4wLi5jUmtMZndKWGhISUdpdW5DM0x4Z2ZRLjYtbmpGQXJBSDhuM01Zcy0zVkJyNGx6NTRsWjNkMTZiNW5objNsaFZoRV9lU0JnU2YzaGlRWDduTFh3cEtJUnhmMXZybGtSRmVPNGxnZ1pzazRRM3JXNjlfbndUUjdBdTdoeUh4NVZRTkdmWG9VVFRmT0k4NGJfaWRzalBwaENyRVExNWdFNHR2TFVXTE5RQTJRUGw5c0lzT2haeW50bHJBTU15UTBZX1JRbG01cktMeEVGd2hMb0xjR0FmaV96dGJCNVJRbGdZSHVIYjFfWkd5OXJMNEVWVWoxeld0VER4SDFBV3laNmlfRGcuS1NzSndKbGo2YllHRFJTcjdQM3VGZw.mnYj8Fu3jzieBZ7Z2TBIDcvk98fJsUis60RSyrEFePKsRo6rz8MdDokcC8b1GeFMrEmyjvYaBuCRIfBzmkR8BA",
    "accessToken": "eyJhbGciOiJIUzUxMiJ9.ZXlKNmFYQWlPaUpFUlVZaUxDSmhiR2NpT2lKa2FYSWlMQ0psYm1NaU9pSkJNVEk0UTBKRExVaFRNalUySW4wLi5jUmtMZndKWGhISUdpdW5DM0x4Z2ZRLjYtbmpGQXJBSDhuM01Zcy0zVkJyNGx6NTRsWjNkMTZiNW5objNsaFZoRV9lU0JnU2YzaGlRWDduTFh3cEtJUnhmMXZybGtSRmVPNGxnZ1pzazRRM3JXNjlfbndUUjdBdTdoeUh4NVZRTkdmWG9VVFRmT0k4NGJfaWRzalBwaENyRVExNWdFNHR2TFVXTE5RQTJRUGw5c0lzT2haeW50bHJBTU15UTBZX1JRbG01cktMeEVGd2hMb0xjR0FmaV96dGJCNVJRbGdZSHVIYjFfWkd5OXJMNEVWVWoxeld0VER4SDFBV3laNmlfRGcuS1NzSndKbGo2YllHRFJTcjdQM3VGZw.mnYj8Fu3jzieBZ7Z2TBIDcvk98fJsUis60RSyrEFePKsRo6rz8MdDokcC8b1GeFMrEmyjvYaBuCRIfBzmkR8BA",
    "sensorsdata2015jssdkchannel": "%7B%22prop%22%3A%7B%22_sa_channel_landing_url%22%3A%22%22%7D%7D",
    "sensorsdata2015jssdkcross": "%7B%22distinct_id%22%3A%22c0db24d8680b49b4a8678ad0a36910e8%22%2C%22first_id%22%3A%221983f6084ec1dd-0648c2f7a46a794-26011951-1327104-1983f6084ed15a5%22%2C%22props%22%3A%7B%22%24latest_traffic_source_type%22%3A%22%E7%9B%B4%E6%8E%A5%E6%B5%81%E9%87%8F%22%2C%22%24latest_search_keyword%22%3A%22%E6%9C%AA%E5%8F%96%E5%88%B0%E5%80%BC_%E7%9B%B4%E6%8E%A5%E6%89%93%E5%BC%80%22%2C%22%24latest_referrer%22%3A%22%22%7D%2C%22identities%22%3A%22eyIkaWRlbnRpdHlfY29va2llX2lkIjoiMTk4M2Y2MDg0ZWMxZGQtMDY0OGMyZjdhNDZhNzk0LTI2MDExOTUxLTEzMjcxMDQtMTk4M2Y2MDg0ZWQxNWE1IiwiJGlkZW50aXR5X2xvZ2luX2lkIjoiYzBkYjI0ZDg2ODBiNDliNGE4Njc4YWQwYTM2OTEwZTgifQ%3D%3D%22%2C%22history_login_id%22%3A%7B%22name%22%3A%22%24identity_login_id%22%2C%22value%22%3A%22c0db24d8680b49b4a8678ad0a36910e8%22%7D%2C%22%24device_id%22%3A%221983f613e1a156a-0bae2da6cdc884-26011951-1327104-1983f613e1b1723%22%7D",
    "acw_tc": "ac11000117534326047683758e00679599479c00600dee3f4908006d898b4e",
    "Hm_lvt_9ea3e7293b7c088e0d2c88874b63e7dd": "1753409995,1753432606",
    "Hm_lpvt_9ea3e7293b7c088e0d2c88874b63e7dd": "1753432606",
    "HMACCOUNT": "29C8E5B0264CE645",
    "creditNo": "%22%22",
    "patentDetailTechnicalSupportShowTime": "show",
    "x-web-ip": "182.143.130.154, 119.23.123.151, 120.78.44.162, 100.121.108.240"
}
url = "https://app.qizhidao.com/qzd-bff-patent/patent/info/record"
data = {
    "pn": "WZIP_14feea8dc3c20cb97f1de11e4233613c",
    "limit": True,
    "statement": "ASPS_EXACT:(\"西南财经大学天府学院\")",
    "subWordSwitch": None,
    "simpleVersion": False
}
data = json.dumps(data, separators=(',', ':'))
response = requests.post(url, headers=headers, cookies=cookies, data=data)


key_index = response.json()["hasUse"]
json_data = response.json()['data1']

# 密钥映射表（来自原始 JS 代码）
AES_KEYS = {
    1: "xc46VoB49X3PGYAg",
    2: "KE3pb84wxqLTZEG3",
    3: "18Lw0OEaBBUwHYNT",
    4: "jxxWWIzvkqEQcZrd",
    5: "40w42rjLEXxYhxRn",
    6: "K6hkD03WNW8N1fPM",
    7: "I8V3IwIhrwNbWxqz",
    8: "3JNNbxAP4zi5oSGA",
    9: "7pUuESQl8aRTFFKK",
    10: "KB4GAHN6M5soB3WV"
}


def aes_ecb_decrypt_base64(cipher_text: str, key_index: int) -> str:
    """
    解密 AES ECB 模式的 base64 密文。
    :param cipher_text: 加密后的 base64 字符串
    :param key_index: 使用的密钥编号（1-10）
    :return: 解密后的明文字符串
    """
    if key_index not in AES_KEYS:
        raise ValueError(f"无效的密钥编号：{key_index}")

    key = AES_KEYS[key_index].encode('utf-8')
    cipher_data = base64.b64decode(cipher_text)
    cipher = AES.new(key, AES.MODE_ECB)
    decrypted = cipher.decrypt(cipher_data)
    # 去除 PKCS7 padding
    pad_len = decrypted[-1]
    return decrypted[:-pad_len].decode('utf-8')





def export_patent_to_excel(data, filename="patent_data.xlsx"):
    """
    将专利数据写入Excel文件。
    参数:
      - data: dict，专利数据
      - filename: str，输出Excel文件名
    """
    def timestamp_to_str(ts):
        """时间戳(毫秒)转日期字符串"""
        try:
            return datetime.fromtimestamp(ts / 1000).strftime("%Y-%m-%d")
        except:
            return str(ts)

    def list_to_str(lst):
        """列表转分号连接字符串"""
        if not lst:
            return ""
        res = []
        for item in lst:
            if isinstance(item, (dict, list)):
                res.append(json.dumps(item, ensure_ascii=False))
            else:
                res.append(str(item))
        return "; ".join(res)

    wb = Workbook()
    ws = wb.active
    ws.title = "专利详情"
    row = 1

    for key, val in data.items():
        # 处理特殊字段格式
        if key in ["app_date", "out_date", "invalid_date"]:
            val = timestamp_to_str(val)
        elif isinstance(val, list):
            val = list_to_str(val)
        elif isinstance(val, dict):
            val = json.dumps(val, ensure_ascii=False)
        elif isinstance(val, bool):
            val = str(val)
        ws.cell(row=row, column=1, value=key)
        ws.cell(row=row, column=2, value=val)
        row += 1

    wb.save(filename)
    print(f"✅ 专利数据已成功导出到文件：{filename}")


if __name__ == '__main__':
    data = aes_ecb_decrypt_base64(json_data, key_index)
    export_patent_to_excel(json.loads(data))

